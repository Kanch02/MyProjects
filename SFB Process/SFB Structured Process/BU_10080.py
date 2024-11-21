import yaml
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import logging
import config
from config import setup_logging
import warnings

warnings.filterwarnings("ignore")

setup_logging()

# -------------------------------------------------------------------------------------------------------------------

def add_additional_columns(today_transaction_file_df, HUB_Source_list, MONTH, YEAR):
    today_transaction_file_df["Month"] = MONTH
    today_transaction_file_df["Year"] = int(YEAR)
    today_transaction_file_df["Business Unit"] = today_transaction_file_df["Business Unit"].astype(int)
    today_transaction_file_df["GL Accountt"] = today_transaction_file_df["GL Account"].astype(int)

    today_transaction_file_df["Revised Journal Source"] = today_transaction_file_df['Journal Source']
    today_transaction_file_df["Journal Source2"] = today_transaction_file_df['Journal Source']

    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Revised Journal Source'] = 'HUB'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Journal Source2'] = 'HUB'
    today_transaction_file_df["Daily Rec Comment"] = "Plug"
    today_transaction_file_df["Master File Status"] = "Plug"

    source_exclude_list = ["9RT"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_exclude_list), 'Revised Journal Source'] = '9RT'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_exclude_list), 'Journal Source2'] = '9RT'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_exclude_list), 'Daily Rec Comment'] = 'Others'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_exclude_list), 'Master File Status'] = 'Minor Diff'

    source_9AQ_list = ["9AQ"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_9AQ_list), 'Revised Journal Source'] = '9AQ'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_9AQ_list), 'Journal Source2'] = '9AQ'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_9AQ_list), 'Daily Rec Comment'] = 'Others'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_9AQ_list), 'Master File Status'] = 'Cleared'

    source_P_list = ["PJE", "PJA", "MAN"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_P_list), 'Daily Rec Comment'] = ''
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(source_P_list), 'Master File Status'] = ''

    return today_transaction_file_df

# -------------------------------------------------------------------------------------------------------------------

def format_amount_columns(dataframe):
    dataframe["Transactional Amount"] = dataframe["Transactional Amount"].str.replace(",", "")
    dataframe["Transactional Amount"] = dataframe["Transactional Amount"].str.replace(")", "")
    dataframe["Transactional Amount"] = dataframe["Transactional Amount"].str.replace("(", "-").astype(float)

    dataframe["Functional Amount"] = dataframe["Functional Amount"].str.replace(",", "")
    dataframe["Functional Amount"] = dataframe["Functional Amount"].str.replace(")", "")
    dataframe["Functional Amount"] = dataframe["Functional Amount"].str.replace("(", "-").astype(float)

    return dataframe

# -------------------------------------------------------------------------------------------------------------------

def main_process_for_nongbp(transaction_file_df, master_file_df, rha_extract_sheet_df, values_for_10080):
    HUB_SOURCE_LIST = values_for_10080['HUB_SOURCE_LIST']
    HUB_SOURCE_LIST = [source.strip() for source in HUB_SOURCE_LIST.split(",")]
    PLATFORM_PROC_DATE = values_for_10080['PLATFORM_PROC_DATE']
    MONTH = values_for_10080['MONTH']
    YEAR = values_for_10080['YEAR']
    CURRENCY = values_for_10080['CURRENCY']
    MASTER_FILE_GL_SHEETNAME = values_for_10080['MASTER_FILE_GL_SHEETNAME']
    MASTER_FILE_RHA_SHEETNAME = values_for_10080['MASTER_FILE_RHA_SHEETNAME']
    RHA_WORK_OF_DATE = values_for_10080['RHA_WORK_OF_DATE']
    RHA_REPORTING_UNIT = values_for_10080['RHA_REPORTING_UNIT']
    RHA_MONTH_DETAIL = values_for_10080['RHA_MONTH_DETAIL']
    REVISED_VALUE_DATE = values_for_10080['REVISED_VALUE_DATE']
    REVISED_WORK_OF_DATE = values_for_10080['REVISED_WORK_OF_DATE']

    logging.info(f"Processing for NON-{CURRENCY} Currency")

    # ---------------------------------------------------
    # From the entire transaction_file_df, we are extracting for today
    today_transaction_file_df = transaction_file_df.loc[
        (transaction_file_df['Platform Proc Date']==PLATFORM_PROC_DATE) & 
        (transaction_file_df['Journal Source'] != 'GLR') &
        (transaction_file_df["Transactional Currency"] != 'GBP')
    ]

    # We want to format the Transactional Amount and Functional Amount to proper float format, so we are calling this function
    today_transaction_file_df = format_amount_columns(today_transaction_file_df)
    # As we know we want extra columns like Jounral Source2 and Revised Journal Source. So we are calling this function to do those changes
    today_transaction_file_df = add_additional_columns(today_transaction_file_df, HUB_SOURCE_LIST, MONTH, YEAR)

    # ---------------------------------------------------

    # Extracting NON-GBP values from RHA FILE
    rha_extract_sheet_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['CURRENCY'] != CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') &
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]
    
    rha_extract_sheet_df["MONTH"] = RHA_MONTH_DETAIL
    rha_extract_sheet_df["OPENING BALANCE"] = ""
    rha_extract_sheet_df["STATUS"] = "Open"
    rha_extract_sheet_df["MONTH.1"] = RHA_MONTH_DETAIL
    rha_extract_sheet_df["YEAR"] = int(YEAR)
    rha_extract_sheet_df["CONCAT"] = ""
    rha_extract_sheet_df["REVISED VALUE DATE"] = REVISED_VALUE_DATE
    rha_extract_sheet_df["REVISED WORK OF DATE"] = REVISED_WORK_OF_DATE

    # Updating the GL sheet
    logging.info(f"Updating the Master GL Sheet in the Master file with Non-GBP data")
    transaction_file_sheet_df = master_file_df [MASTER_FILE_GL_SHEETNAME]
    aligned_today_transaction_file_df = today_transaction_file_df.reindex(columns=transaction_file_sheet_df.columns)
    updated_transaction_file_sheet_df = pd.concat([transaction_file_sheet_df, aligned_today_transaction_file_df], ignore_index=True) 
    master_file_df[MASTER_FILE_GL_SHEETNAME] = updated_transaction_file_sheet_df

    # Updating the RHA Sheet
    logging.info(f"Updating the RHA Sheet in the Master file with with Non-GBP data")
    rha_master_sheet_df = master_file_df[MASTER_FILE_RHA_SHEETNAME]
    rha_master_sheet_df.columns = [col.upper() for col in rha_master_sheet_df.columns]
    aligned_rha_master_sheet_df = rha_extract_sheet_df.reindex(columns=rha_master_sheet_df.columns)
    updated_rha_master_sheet_df = pd.concat([rha_master_sheet_df, aligned_rha_master_sheet_df], ignore_index=True)
    master_file_df[MASTER_FILE_RHA_SHEETNAME] = updated_rha_master_sheet_df

    # ---------------------------------------------------

    logging.info(f"NON-{CURRENCY} Values updation happened in the Master file successfully !!!")
    logging.info(f"Process completed for NON-{CURRENCY} Currency")

# -------------------------------------------------------------------------------------------------------------------

def main_process_for_gbp(transaction_file_df, today_plug_file_df, master_file_df, rha_extract_sheet_df, values_for_10080):
    HUB_SOURCE_LIST = values_for_10080['HUB_SOURCE_LIST']
    HUB_SOURCE_LIST = [source.strip() for source in HUB_SOURCE_LIST.split(",")]
    PLATFORM_PROC_DATE = values_for_10080['PLATFORM_PROC_DATE']
    MONTH = values_for_10080['MONTH']
    YEAR = values_for_10080['YEAR']
    CURRENCY = values_for_10080['CURRENCY']
    MASTER_FILE_PLUG_SHEETNAME = values_for_10080['MASTER_FILE_PLUG_SHEETNAME']
    MASTER_FILE_GL_SHEETNAME = values_for_10080['MASTER_FILE_GL_SHEETNAME']
    MASTER_FILE_RHA_SHEETNAME = values_for_10080['MASTER_FILE_RHA_SHEETNAME']
    RHA_WORK_OF_DATE = values_for_10080['RHA_WORK_OF_DATE']
    RHA_REPORTING_UNIT = values_for_10080['RHA_REPORTING_UNIT']
    RHA_MONTH_DETAIL = values_for_10080['RHA_MONTH_DETAIL']
    REVISED_VALUE_DATE = values_for_10080['REVISED_VALUE_DATE']
    REVISED_WORK_OF_DATE = values_for_10080['REVISED_WORK_OF_DATE']

    # ---------------------------------------------------
    logging.info(f"Processing for {CURRENCY} Currency")

    # From the entire transaction_file_df, we are extracting for today
    today_transaction_file_df = transaction_file_df.loc[
        (transaction_file_df['Platform Proc Date']==PLATFORM_PROC_DATE) & 
        (transaction_file_df['Journal Source'] != 'GLR') &
        (transaction_file_df["Transactional Currency"] == 'GBP')
    ]

    # We want to format the Transactional Amount and Functional Amount to proper float format, so we are calling this function
    today_transaction_file_df = format_amount_columns(today_transaction_file_df)
    # As we know we want extra columns like Jounral Source2 and Revised Journal Source. So we are calling this function to do those changes
    today_transaction_file_df = add_additional_columns(today_transaction_file_df, HUB_SOURCE_LIST, MONTH, YEAR)

    # ---------------------------------------------------

    # From the entire today_plug_file_df, we are extracting ppd dataframe
    ppd_plug_file_df = today_plug_file_df.loc[
        (today_plug_file_df['Platform Proc Date'] == PLATFORM_PROC_DATE) &
        (today_plug_file_df['Journal Source'] != 'GLR')
    ]

    # We want to format the Transactional Amount and Functional Amount to proper float format, so we are calling this function
    ppd_plug_file_df = format_amount_columns(ppd_plug_file_df)

    # ---------------------------------------------------

    # Defining rha_tobe_added_df dataframe
    rha_tobe_added_df = pd.DataFrame()

    # ---------------------------------------------------

    # FROM HERE STARTING THE MAIN PROCESS
    transaction_amount_sum = ppd_plug_file_df["Transactional Amount"].sum()
    logging.info(f"Total Transactional Amount sum from Plug File = {transaction_amount_sum}")

    total_gl_sum = today_transaction_file_df["Functional Amount"].sum()
    logging.info(f"Total Functional Amount sum for all HUB Sources from Transactional File = {total_gl_sum}")

    # BELOW WE WILL BE LOOPING THROUGH ALL THE HUB SOURCES AND REPEATING THE STEPS
    HUB_SOURCE_LIST = [source.strip() for source in HUB_SOURCE_LIST.split(",")]
    for source in HUB_SOURCE_LIST:
        logging.info(f"FOR SOURCE = {source}")
        # PPD details
        journal_source_df = ppd_plug_file_df.loc[(ppd_plug_file_df["Journal Source"] == source)]
        source_sum = journal_source_df["Transactional Amount"].sum()
        logging.info(f"Total transactional sum from plug file for source {source} = {source_sum}")

        # Transactional File Details
        master_gl_source_df = today_transaction_file_df.loc[
            (today_transaction_file_df["Transactional Currency"] == CURRENCY) & 
            (today_transaction_file_df["Journal Source"] == source) 
        ]
        total_gl = master_gl_source_df["Functional Amount"].sum()
        logging.info(f"Functional Amount sum from Transactional file for {source} = {total_gl}")

        # Difference Calculation
        source_difference_amount = round((total_gl + source_sum), 2)
        logging.info(f"Difference in {source}: {source_difference_amount}")

        # Checking in RHA file
        difference_source_df = pd.DataFrame()
        logging.info(f"Checking for source {source}:")
        if(source_difference_amount != 0):
            rha_extract_source_df = rha_extract_sheet_df.loc[
                (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
                (rha_extract_sheet_df['SSID'] == source) &
                (rha_extract_sheet_df['CURRENCY'] == CURRENCY) &
                (rha_extract_sheet_df['T/R'] == 'T') & 
                (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
            ]
            particular_rha_difference_source = rha_extract_source_df.loc[(rha_extract_source_df['AMOUNT'] == source_difference_amount)]
            if(particular_rha_difference_source.empty):
                total_amount_rha = round(rha_extract_source_df['AMOUNT'].sum(), 2)
                if(total_amount_rha == source_difference_amount):
                    difference_source_df = rha_extract_source_df
                    logging.info(f"Found the difference in RHA file for source {source}")
                else:
                    logging.info(f"Didnt found any matching RHA file for source {source}")
            else:
                difference_source_df = particular_rha_difference_source
                logging.info(f"Found the difference in RHA file for source {source}")
        else:
            logging.info(f"Difference for Source {source} = 0")

        if not difference_source_df.empty: # Note: Here for all the source loc statement should be checked                                           <<<---------------------
            # First line got updated
            today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == source) & today_transaction_file_df["Functional Amount"])]
            today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == source) & today_transaction_file_df["Functional Amount"])]
            today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == source) & today_transaction_file_df["Functional Amount"])]

            # Second line updated
            copy_line_source_df = today_transaction_file_df.loc[(today_transaction_file_df["Journal Source"] == source) & (today_transaction_file_df["Functional Amount"])]
            copy_line_source_df["Functional Amount"] = ""
            copy_line_source_df["Daily Rec Comment"] = "RHA"
            copy_line_source_df["Master File Status"] = "RHA"
            copy_line_source_df["Transactional Amount"] = source_difference_amount
            # Adding the RHA reject into the RHA sheet inside Master file
            rha_tobe_added_df = pd.concat([rha_tobe_added_df, difference_source_df], ignore_index=True)

    # End of for loop
    # ---------------------------------------------------

    # Adding all the extra columns to the rha_tobe_added dataframe
    if (not rha_tobe_added_df.empty):
        rha_tobe_added_df["MONTH"] = RHA_MONTH_DETAIL
        rha_tobe_added_df["OPENING BALANCE"] = ""
        rha_tobe_added_df["STATUS"] = "Open"
        rha_tobe_added_df["MONTH.1"] = RHA_MONTH_DETAIL
        rha_tobe_added_df["YEAR"] = int(YEAR)
        rha_tobe_added_df["CONCAT"] = ""
        rha_tobe_added_df["REVISED VALUE DATE"] = REVISED_VALUE_DATE
        rha_tobe_added_df["REVISED WORK OF DATE"] = REVISED_WORK_OF_DATE

    # Updating the GL sheet
    logging.info(f"Updating the Master GL Sheet in the Master file with all the transaction file data and the RHA rejects.")
    transaction_file_sheet_df = master_file_df [MASTER_FILE_GL_SHEETNAME]
    aligned_today_transaction_file_df = today_transaction_file_df.reindex(columns=transaction_file_sheet_df.columns)
    updated_transaction_file_sheet_df = pd.concat([transaction_file_sheet_df, aligned_today_transaction_file_df], ignore_index=True) 
    master_file_df[MASTER_FILE_GL_SHEETNAME] = updated_transaction_file_sheet_df

    # Updating the RHA Sheet
    logging.info(f"Updating the RHA Sheet in the Master file with all the RHA rejects.")
    rha_master_sheet_df = master_file_df[MASTER_FILE_RHA_SHEETNAME]
    rha_master_sheet_df.columns = [col.upper() for col in rha_master_sheet_df.columns]
    aligned_rha_master_sheet_df = rha_tobe_added_df.reindex(columns=rha_master_sheet_df.columns)
    updated_rha_master_sheet_df = pd.concat([rha_master_sheet_df, aligned_rha_master_sheet_df], ignore_index=True)
    master_file_df[MASTER_FILE_RHA_SHEETNAME] = updated_rha_master_sheet_df

    # Updating the Plug Sheet
    logging.info(f"Updating the Plug Sheet in the Master file with all the Today plug file data.")
    plug_sheet_masterfile_df = master_file_df[MASTER_FILE_PLUG_SHEETNAME]
    aligned_plugdata_ppd_df = ppd_plug_file_df.reindex(columns=plug_sheet_masterfile_df.columns)
    updated_plug_sheet_masterfile_df = pd.concat([plug_sheet_masterfile_df, aligned_plugdata_ppd_df], ignore_index=True)
    master_file_df[MASTER_FILE_PLUG_SHEETNAME] = updated_plug_sheet_masterfile_df

    logging.info(f"{CURRENCY} Values updation happened in the Master file successfully !!!")
    logging.info(f"Process completed for {CURRENCY} Currency")

# -------------------------------------------------------------------------------------------------------------------

def sfb_process_10080():
    values_for_10080 = config.values_of_10080

    # Extracting all the FILE PATH values from the variable values_for_10080
    TRANSACTION_FILE_PATH = values_for_10080['TRANSACTION_FILE_PATH']
    PLUG_FILE_PATH = values_for_10080['PLUG_FILE_PATH']
    MASTER_FILE_PATH = values_for_10080['MASTER_FILE_PATH']
    RHA_FILE_PATH = values_for_10080['RHA_FILE_PATH']
    

    # 1) Reading the Transactional file and removing the = and " from the column name of dataframe
    headerrow = 46
    transaction_file = TRANSACTION_FILE_PATH
    logging.info(f"Reading today's transaction file from path: {os.path.basename(transaction_file)}")
    transaction_file_df = pd.read_csv(transaction_file, header = headerrow, skip_blank_lines = False)
    transaction_file_df = transaction_file_df.replace(["=",'"'],"", regex = True)
    transaction_file_df.columns = transaction_file_df.columns.str.replace("=","")
    transaction_file_df.columns = transaction_file_df.columns.str.replace('"','')

    # 2) Reading the Plug File and removing the = and " from the column name of dataframe
    headerrow = 46
    plug_file = PLUG_FILE_PATH
    logging.info(f"Reading the plug file from path: {os.path.basename(plug_file)}")
    today_plug_file_df = pd.read_csv(plug_file, header = headerrow, skip_blank_lines = False)
    today_plug_file_df = today_plug_file_df.replace(["=",'"'],"", regex = True)
    today_plug_file_df.columns = today_plug_file_df.columns.str.replace("=","")
    today_plug_file_df.columns = today_plug_file_df.columns.str.replace('"','')

    # 3) Reading the Master File
    master_file = MASTER_FILE_PATH
    master_file_df = pd.read_excel(master_file, sheet_name=None)

    # 4) Reading RHA File and making the columns name to UPPER CASE
    rha_file = RHA_FILE_PATH
    logging.info(f"Reading the RHA File to find the rejects exist or not from file path: {os.path.basename(rha_file)} ")
    rha_file_df = pd.read_excel(rha_file, sheet_name=None)
    rha_sheetname = list(rha_file_df.keys())[0]
    rha_extract_sheet_df = rha_file_df[rha_sheetname]
    rha_extract_sheet_df.columns = [col.upper() for col in rha_extract_sheet_df.columns]

    # Now calling the respective function to do their jobs
    main_process_for_gbp(transaction_file_df, today_plug_file_df, master_file_df, rha_extract_sheet_df, values_for_10080)
    main_process_for_nongbp(transaction_file_df, master_file_df, rha_extract_sheet_df, values_for_10080)


    # Once final updating happened now adding those values to the excel file
    with pd. ExcelWriter(master_file, engine='openpyxl') as writer:
        for sheet_name, data in master_file_df.items(): 
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    # Formatting the SFB File
    master_file_df = pd.read_excel(master_file, sheet_name=None)
    wb = load_workbook(master_file)
    yellow_fill = PatternFill(start_color='E8FF00', end_color='E8FF00', fill_type='solid')

    for sheet_name, data in master_file_df.items():
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = yellow_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    wb.save(master_file)

# -------------------------------------------------------------------------------------------------------------------

