import yaml
import os
import pandas as pd # type: ignore
import openpyxl # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import PatternFill # type: ignore
from datetime import datetime
import logging
import warnings

warnings.filterwarnings("ignore")

log_folder = ".\\Logs"
os.makedirs(log_folder, exist_ok=True)
log_file = log_folder + "\\SFB_PROCESS_FLOW_" + datetime.now().strftime("%m%d%Y-%H%M%S") + ".log"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)-7s | %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StramHandler()
    ]
)

try:
    logging.info("Starting the SFB Process !!!")
    logging.info("-"*30)
    logging.info("Processing for GBP Currency")

    # ----------------------------------------------------------------------------
    config_filepath = "SFB Config.yaml"
    logging.info(f"Retriving the values from the file: {os.path.basename(config_filepath)}")

    with open(config_filepath, 'r') as file:
        config = yaml.safe_load(file)

    TRANSACTION_FILE_PATH = config['TRANSACTION_FILE_PATH']
    PLATFORM_PROC_DATE = config['PLATFORM_PROC_DATE']
    MONTH = config['MONTH']
    YEAR = config['YEAR']
    CURRENCY = config['CURRENCY']
    PLUG_FILE_PATH = config['PLUG_FILE_PATH']
    MASTER_FILE_PATH = config['MASTER_FILE_PATH']
    MASTER_FILE_PLUG_SHEETNAME = config['MASTER_FILE_PLUG_SHEETNAME']
    MASTER_FILE_GL_SHEETNAME = config['MASTER_FILE_GL_SHEETNAME']
    MASTER_FILE_RHA_SHEETNAME = config['MASTER_FILE_RHA_SHEETNAME']

    RHA_FILE_PATH = config['RHA_FILE_PATH']
    RHA_WORK_OF_DATE = config['RHA_WORK_OF_DATE']
    RHA_REPORTING_UNIT = config['RHA_REPORTING_UNIT']
    
    RHA_MONTH_DETAIL = config['RHA_MONTH_DETAIL']
    REVISED_VALUE_DATE = config['REVISED_VALUE_DATE']
    REVISED_WORK_OF_DATE = config['REVISED_WORK_OF_DATE']

    logging.info(f"Retriving completed from config file")
    # ----------------------------------------------------------------------------

    headerrow = 46
    transaction_file = TRANSACTION_FILE_PATH
    
    logging.info(f"Reading the today transaction file from path: {os.path.basename(transaction_file)}")

    transaction_file_df = pd.read_csv(transaction_file, header = headerrow, skip_blank_lines = False)
    transaction_file_df = transaction_file_df.replace(["=",'"'],"", regex = True)
    transaction_file_df.columns = transaction_file_df.columns.str.replace("=","")
    transaction_file_df.columns = transaction_file_df.columns.str.replace('"','')

    # Note: Here the line has to be completed                                                               <<<---------------------
    today_transaction_file_df = transaction_file_df.loc[
        (transaction_file_df['Platform Proc Date']==PLATFORM_PROC_DATE) & 
        (transaction_file_df['Journal Source'] != 'GLR') &
        (transaction_file_df["Transactional Currency"] == 'GBP')
    ]

    today_transaction_file_df["Transactional Amount"] = today_transaction_file_df["Transactional Amount"].str.replace(",", "")
    today_transaction_file_df["Transactional Amount"] = today_transaction_file_df["Transactional Amount"].str.replace(")", "")
    today_transaction_file_df["Transactional Amount"] = today_transaction_file_df["Transactional Amount"].str.replace("(", "-").astype(float)

    today_transaction_file_df["Functional Amount"] = today_transaction_file_df["Functional Amount"].str.replace(",", "")
    today_transaction_file_df["Functional Amount"] = today_transaction_file_df["Functional Amount"].str.replace(")", "")
    today_transaction_file_df["Functional Amount"] = today_transaction_file_df["Functional Amount"].str.replace("(", "-").astype(float)

    today_transaction_file_df["Revised Journal Source"] = today_transaction_file_df['Journal Source']
    today_transaction_file_df["Journal Source2"] = today_transaction_file_df['Journal Source']

    HUB_Source_list = ["91G", "91P", "91I", "91E", "91T", "91D"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Revised Journal Source'] = 'HUB'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Journal Source2'] = 'HUB'
    today_transaction_file_df["Daily Rec Comment"] = "Plug"
    today_transaction_file_df["Master File Status"] = "Plug"

    logging.info(f"Modified the Revised Journal Source and Journal Source2 to HUB for all the HUB Sources: {HUB_Source_list}")

    source_exclude_list = ["9RT"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Revised Journal Source'] = '9RT'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Journal Source2'] = '9RT'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Daily Rec Comment'] = 'Others'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Master File Status'] = 'Minor Diff'

    source_9AQ_list = ["9AQ"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Revised Journal Source'] = '9AQ'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Journal Source2'] = '9AQ'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Daily Rec Comment'] = 'Others'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Master File Status'] = 'Cleared'

    source_P_list = ["PJE", "PJA", "MAN"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Daily Rec Comment'] = ''
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Master File Status'] = ''

    # ---------------------------------------------------------------------

    headerrow = 46
    plug_file = PLUG_FILE_PATH

    logging.info(f"Reading the plug file from path: {os.path.basename(plug_file)}")

    today_plug_file_df = pd.read_csv(plug_file, header = headerrow, skip_blank_lines = False)
    today_plug_file_df = today_plug_file_df.replace(["=",'"'],"", regex = True)
    today_plug_file_df.columns = today_plug_file_df.columns.str.replace("=","")
    today_plug_file_df.columns = today_plug_file_df.columns.str.replace('"','')

    # Note: Here the line has to be completed                                                               <<<---------------------
    ppd_plug_file_df = today_plug_file_df.loc[
        (today_plug_file_df['Platform Proc Date'] == PLATFORM_PROC_DATE) &
        (today_plug_file_df['Journal Source'] != 'GLR')
    ]

    ppd_plug_file_df["Transactional Amount"] = ppd_plug_file_df["Transactional Amount"].str.replace(",", "")
    ppd_plug_file_df["Transactional Amount"] = ppd_plug_file_df["Transactional Amount"].str.replace(")", "")
    ppd_plug_file_df["Transactional Amount"] = ppd_plug_file_df["Transactional Amount"].str.replace("(", "-").astype(float)

    ppd_plug_file_df["Functional Amount"] = ppd_plug_file_df["Functional Amount"].str.replace(",", "")
    ppd_plug_file_df["Functional Amount"] = ppd_plug_file_df["Functional Amount"].str.replace(")", "")
    ppd_plug_file_df["Functional Amount"] = ppd_plug_file_df["Functional Amount"].str.replace("(", "-").astype(float)

    transaction_amount_sum = ppd_plug_file_df["Transactional Amount"].sum()
    logging.info(f"Total Transactional Amount sum from Plug File: {transaction_amount_sum}")
    
    logging.info("Below values are from Plug file:")
    
    #for source 91G
    journal_source_91G_df = ppd_plug_file_df.loc[(ppd_plug_file_df["Journal Source"] == "91G")]
    source_sum91G = journal_source_91G_df["Transactional Amount"].sum()
    logging.info(f"Total transactions sum for source 91G = {source_sum91G}")

    #for source 91P
    journal_source_91P_df = ppd_plug_file_df.loc[(ppd_plug_file_df["Journal Source"] == "91P")]
    source_sum91P = journal_source_91P_df["Transactional Amount"].sum()
    logging.info(f"Total transactions sum for source 91P = {source_sum91P}")

    #for source 91D
    journal_source_91D_df = ppd_plug_file_df.loc[(ppd_plug_file_df["Journal Source"] == "91D")]
    source_sum91D = journal_source_91D_df["Transactional Amount"].sum()
    logging.info(f"Total transactions sum for source 91D = {source_sum91D}")

    #for source 91E
    journal_source_91E_df = ppd_plug_file_df.loc[(ppd_plug_file_df["Journal Source"] == "91E")]
    source_sum91E = journal_source_91E_df["Transactional Amount"].sum()
    logging.info(f"Total transactions sum for source 91E = {source_sum91E}")

    #for source 91I
    journal_source_91I_df = ppd_plug_file_df.loc[(ppd_plug_file_df["Journal Source"] == "91I")]
    source_sum91I = journal_source_91I_df["Transactional Amount"].sum()
    logging.info(f"Total transactions sum for source 91I = {source_sum91I}")

    #for source 91T
    journal_source_91T_df = ppd_plug_file_df.loc[(ppd_plug_file_df["Journal Source"] == "91T")]
    source_sum91T = journal_source_91T_df["Transactional Amount"].sum()
    logging.info(f"Total transactions sum for source 91T = {source_sum91T}")


    logging.info(f"Adding the today's plug file data into Masterfile's plug sheet")
    master_file = MASTER_FILE_PATH
    master_file_df = pd.read_excel(master_file, sheet_name=None)
    plug_masterfile_sheetname = MASTER_FILE_PLUG_SHEETNAME
    plug_sheet_masterfile_df = master_file_df[plug_masterfile_sheetname]
    aligned_plugdata_ppd_df = ppd_plug_file_df.reindex(columns=plug_sheet_masterfile_df.columns)
    updated_plug_sheet_masterfile_df = pd.concat([plug_sheet_masterfile_df, aligned_plugdata_ppd_df], ignore_index=True)
    master_file_df[plug_masterfile_sheetname] = updated_plug_sheet_masterfile_df

    with pd.ExcelWriter(master_file, engine='openpyxl') as writer:
        for sheet_name, data in master_file_df.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)
    
    logging.info(f"Today's plug file data added successfully into Masterfile's plug sheet !!!")

    # ---------------------------------------------------------------------

    logging.info(f"Below are values from Transaction File:")

    total_gl_sum = today_transaction_file_df["Functional Amount"].sum()
    logging.info(f"Total Functional Amount sum for all HUB Sources = {total_gl_sum}")

    # Note: Here for all the source loc statement should be checked                                                  <<<---------------------
    master_gl_source91G_df = today_transaction_file_df.loc[
        (today_transaction_file_df["Transactional Currency"] == CURRENCY) & 
        (today_transaction_file_df["Journal Source"] == "91G") 
    ]
    total_gl_91G = master_gl_source91G_df["Functional Amount"].sum()
    logging.info(f"Functional Amount sum of 91G = {total_gl_91G}")

    master_gl_source91P_df = today_transaction_file_df.loc[
        (today_transaction_file_df["Transactional Currency"] == CURRENCY) & 
        (today_transaction_file_df["Journal Source"] == "91P") 
    ]
    total_gl_91P = master_gl_source91P_df["Functional Amount"].sum()
    logging.info(f"Functional Amount sum of 91P = {total_gl_91P}")

    master_gl_source91D_df = today_transaction_file_df.loc[
        (today_transaction_file_df["Transactional Currency"] == CURRENCY) & 
        (today_transaction_file_df["Journal Source"] == "91D") 
    ]
    total_gl_91D = master_gl_source91D_df["Functional Amount"].sum()
    logging.info(f"Functional Amount sum of 91D = {total_gl_91D}")

    master_gl_source91E_df = today_transaction_file_df.loc[
        (today_transaction_file_df["Transactional Currency"] == CURRENCY) & 
        (today_transaction_file_df["Journal Source"] == "91E") 
    ]
    total_gl_91E = master_gl_source91E_df["Functional Amount"].sum()
    logging.info(f"Functional Amount sum of 91E = {total_gl_91E}")

    master_gl_source91I_df = today_transaction_file_df.loc[
        (today_transaction_file_df["Transactional Currency"] == CURRENCY) & 
        (today_transaction_file_df["Journal Source"] == "91I") 
    ]
    total_gl_91I = master_gl_source91I_df["Functional Amount"].sum()
    logging.info(f"Functional Amount sum of 91I = {total_gl_91I}")

    master_gl_source91T_df = today_transaction_file_df.loc[
        (today_transaction_file_df["Transactional Currency"] == CURRENCY) & 
        (today_transaction_file_df["Journal Source"] == "91T") 
    ]
    total_gl_91T = master_gl_source91T_df["Functional Amount"].sum()
    logging.info(f"Functional Amount sum of 91T = {total_gl_91T}")

    # ---------------------------------------------------------------------

    logging.info(f"Calculating the Difference for all the HUB Sources")

    source_91G_difference_amount = round((total_gl_91G + source_sum91G), 2)
    logging.info(f"Difference in 91G: {source_91G_difference_amount}")

    source_91P_difference_amount = round((total_gl_91P + source_sum91P), 2)
    logging.info(f"Difference in 91P: {source_91P_difference_amount}")

    source_91D_difference_amount = round((total_gl_91D + source_sum91D), 2)
    logging.info(f"Difference in 91D: {source_91D_difference_amount}")

    source_91E_difference_amount = round((total_gl_91E + source_sum91E), 2)
    logging.info(f"Difference in 91E: {source_91E_difference_amount}")

    source_91I_difference_amount = round((total_gl_91I + source_sum91I), 2)
    logging.info(f"Difference in 91I: {source_91I_difference_amount}")

    source_91T_difference_amount = round((total_gl_91T + source_sum91T), 2)
    logging.info(f"Difference in 91T: {source_91T_difference_amount}")

    # ---------------------------------------------------------------------

    rha_file = RHA_FILE_PATH
    logging.info(f"Reading the RHA File to find the rejects exist or not from file path {os.path.basename(rha_file)} ")

    rha_file_df = pd.read_excel(rha_file, sheet_name=None)
    rha_sheetname = list(rha_file_df.keys())[0]
    rha_extract_sheet_df = rha_file_df[rha_sheetname]
    rha_extract_sheet_df.columns = [col.upper() for col in rha_extract_sheet_df.columns]
    
    difference_source91G_df = pd.DataFrame() 
    difference_source91P_df = pd.DataFrame()
    difference_source91D_df = pd.DataFrame()
    difference_source91E_df = pd.DataFrame() 
    difference_source91I_df = pd.DataFrame()
    difference_source91T_df = pd.DataFrame()

    # Note: Here for all the source loc statement should be checked                                                  <<<---------------------

    # Doing for Source 91G
    logging.info(f"Checking for source 91G:")
    if(source_91G_difference_amount != 0):
        rha_extract_source91G_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['SSID'] == "91G") &
            (rha_extract_sheet_df['CURRENCY'] == CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') & 
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]
        particular_rha_difference_source_91G = rha_extract_source91G_df.loc[(rha_extract_source91G_df['AMOUNT'] == source_91G_difference_amount)]
        if(particular_rha_difference_source_91G.empty):
            total_amount_rha_91G = round(rha_extract_source91G_df['AMOUNT'].sum(), 2)
            if(total_amount_rha_91G == source_91G_difference_amount):
                difference_source91G_df = rha_extract_source91G_df
                logging.info(f"Found the difference in RHA file for source 91G")
            else:
                logging.info(f"Didnt found any matching RHA file for source 91G")
        else:
            difference_source91G_df = particular_rha_difference_source_91G
            logging.info("Found the difference in RHA file for source 91G")
    else:
        logging.info("Difference for Source 91G = 0")

    # Doing for Source 91P
    logging.info(f"Checking for source 91P:")
    if(source_91P_difference_amount != 0):
        rha_extract_source91P_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['SSID'] == "91P") &
            (rha_extract_sheet_df['CURRENCY'] == CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') &
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]
        particular_rha_difference_source_91P = rha_extract_source91P_df.loc[(rha_extract_source91P_df['AMOUNT'] == source_91P_difference_amount)]
        if(particular_rha_difference_source_91P.empty):
            total_amount_rha_91P = round(rha_extract_source91P_df['AMOUNT'].sum(), 2)
            if(total_amount_rha_91P == source_91P_difference_amount):
                difference_source91P_df = rha_extract_source91P_df
                logging.info(f"Found the difference in RHA file for source 91P")
            else:
                logging.info(f"Didnt found any matching RHA file for source 91P")
        else:
            difference_source91P_df = particular_rha_difference_source_91P
            logging.info("Found the difference in RHA file for source 91P")
    else:
        logging.info("Difference for Source 91P = 0")

    # Doing for Source 91D
    logging.info(f"Checking for source 91D:")
    if(source_91D_difference_amount != 0):
        rha_extract_source91D_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['SSID'] == "91D") &
            (rha_extract_sheet_df['CURRENCY'] == CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') &
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]
        particular_rha_difference_source_91D = rha_extract_source91D_df.loc[(rha_extract_source91D_df['AMOUNT'] == source_91D_difference_amount)]
        if(particular_rha_difference_source_91D.empty):
            total_amount_rha_91D = round(rha_extract_source91D_df['AMOUNT'].sum(), 2)
            if(total_amount_rha_91D == source_91D_difference_amount):
                difference_source91D_df = rha_extract_source91D_df
                logging.info(f"Found the difference in RHA file for source 91D")
            else:
                logging.info(f"Didnt found any matching RHA file for source 91D")
        else:
            difference_source91D_df = particular_rha_difference_source_91D
            logging.info("Found the difference in RHA file for source 91D")
    else:
        logging.info("Difference for Source 91D = 0")

    # Doing for Source 91E
    logging.info(f"Checking for source 91E:")
    if(source_91E_difference_amount != 0):
        rha_extract_source91E_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['SSID'] == "91E") &
            (rha_extract_sheet_df['CURRENCY'] == CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') &
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]
        particular_rha_difference_source_91E = rha_extract_source91E_df.loc[(rha_extract_source91E_df['AMOUNT'] == source_91E_difference_amount)]
        if(particular_rha_difference_source_91E.empty):
            total_amount_rha_91E = round(rha_extract_source91E_df['AMOUNT'].sum(), 2)
            if(total_amount_rha_91E == source_91E_difference_amount):
                difference_source91E_df = rha_extract_source91E_df
                logging.info(f"Found the difference in RHA file for source 91E")
            else:
                logging.info(f"Didnt found any matching RHA file for source 91E")
        else:
            difference_source91E_df = particular_rha_difference_source_91E
            logging.info("Found the difference in RHA file for source 91E")
    else:
        logging.info("Difference for Source 91E = 0")

    # Doing for Source 91I
    logging.info(f"Checking for source 91I:")
    if(source_91I_difference_amount != 0):
        rha_extract_source91I_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['SSID'] == "91I") &
            (rha_extract_sheet_df['CURRENCY'] == CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') &
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]
        particular_rha_difference_source_91I = rha_extract_source91I_df.loc[(rha_extract_source91I_df['AMOUNT'] == source_91I_difference_amount)]
        if(particular_rha_difference_source_91I.empty):
            total_amount_rha_91I = round(rha_extract_source91I_df['AMOUNT'].sum(), 2)
            if(total_amount_rha_91I == source_91I_difference_amount):
                difference_source91I_df = rha_extract_source91I_df
                logging.info(f"Found the difference in RHA file for source 91I")
            else:
                logging.info(f"Didnt found any matching RHA file for source 91I")
        else:
            difference_source91I_df = particular_rha_difference_source_91I
            logging.info("Found the difference in RHA file for source 91I")
    else:
        logging.info("Difference for Source 91I = 0")

    # Doing for Source 91T
    logging.info(f"Checking for source 91T:")
    if(source_91T_difference_amount != 0):
        rha_extract_source91T_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['SSID'] == "91T") &
            (rha_extract_sheet_df['CURRENCY'] == CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') &
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]
        particular_rha_difference_source_91T = rha_extract_source91T_df.loc[(rha_extract_source91T_df['AMOUNT'] == source_91T_difference_amount)]
        if(particular_rha_difference_source_91T.empty):
            total_amount_rha_91T = round(rha_extract_source91T_df['AMOUNT'].sum(), 2)
            if(total_amount_rha_91T == source_91T_difference_amount):
                difference_source91T_df = rha_extract_source91T_df
                logging.info(f"Found the difference in RHA file for source 91T")
            else:
                logging.info(f"Didnt found any matching RHA file for source 91T")
        else:
            difference_source91T_df = particular_rha_difference_source_91T
            logging.info("Found the difference in RHA file for source 91T")
    else:
        logging.info("Difference for Source 91T = 0")

    # ------------------------------------------------------------------

    # Creating an empty dataframe to store all the rha reject which has to be added into the rha sheet inside master file
    rha_tobe_added_df = pd.DataFrame()

    # Note: Here for all the source loc statement should be checked                                                  <<<---------------------

    # Adding for source 91G
    if not difference_source91G_df.empty:
        # First line got updated
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91G") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91G") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91G") & today_transaction_file_df["Functional Amount"])]

        # Second line updated
        copy_line_source91G_df = today_transaction_file_df.loc[(today_transaction_file_df["Journal Source"] == "91G") & (today_transaction_file_df["Functional Amount"])]
        copy_line_source91G_df["Functional Amount"] = ""
        copy_line_source91G_df["Daily Rec Comment"] = "RHA"
        copy_line_source91G_df["Master File Status"] = "RHA"
        copy_line_source91G_df["Transactional Amount"] = source_91G_difference_amount
        # Adding the RHA reject into the RHA sheet inside Master file
        rha_tobe_added_df = pd.concat([rha_tobe_added_df, difference_source91G_df], ignore_index=True)
    # --------------------------
    
    # Adding for source 91P
    if not difference_source91P_df.empty:
        # First line got updated
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91P") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91P") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91P") & today_transaction_file_df["Functional Amount"])]

        # Second line updated
        copy_line_source91P_df = today_transaction_file_df.loc[(today_transaction_file_df["Journal Source"] == "91P") & (today_transaction_file_df["Functional Amount"])]
        copy_line_source91P_df["Functional Amount"] = ""
        copy_line_source91P_df["Daily Rec Comment"] = "RHA"
        copy_line_source91P_df["Master File Status"] = "RHA"
        copy_line_source91P_df["Transactional Amount"] = source_91P_difference_amount
        # Adding the RHA reject into the RHA sheet inside Master file
        rha_tobe_added_df = pd.concat([rha_tobe_added_df, difference_source91P_df], ignore_index=True)
    # --------------------------

    # Adding for source 91D
    if not difference_source91D_df.empty:
        # First line got updated
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91D") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91D") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91D") & today_transaction_file_df["Functional Amount"])]

        # Second line updated
        copy_line_source91D_df = today_transaction_file_df.loc[(today_transaction_file_df["Journal Source"] == "91D") & (today_transaction_file_df["Functional Amount"])]
        copy_line_source91D_df["Functional Amount"] = ""
        copy_line_source91D_df["Daily Rec Comment"] = "RHA"
        copy_line_source91D_df["Master File Status"] = "RHA"
        copy_line_source91D_df["Transactional Amount"] = source_91D_difference_amount
        # Adding the RHA reject into the RHA sheet inside Master file
        rha_tobe_added_df = pd.concat([rha_tobe_added_df, difference_source91D_df], ignore_index=True)
    # --------------------------

    # Adding for source 91E
    if not difference_source91E_df.empty:
        # First line got updated
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91E") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91E") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91E") & today_transaction_file_df["Functional Amount"])]

        # Second line updated
        copy_line_source91E_df = today_transaction_file_df.loc[(today_transaction_file_df["Journal Source"] == "91E") & (today_transaction_file_df["Functional Amount"])]
        copy_line_source91E_df["Functional Amount"] = ""
        copy_line_source91E_df["Daily Rec Comment"] = "RHA"
        copy_line_source91E_df["Master File Status"] = "RHA"
        copy_line_source91E_df["Transactional Amount"] = source_91E_difference_amount
        # Adding the RHA reject into the RHA sheet inside Master file
        rha_tobe_added_df = pd.concat([rha_tobe_added_df, difference_source91E_df], ignore_index=True)
    # --------------------------

    # Adding for source 91I
    if not difference_source91I_df.empty:
        # First line got updated
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91I") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91I") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91I") & today_transaction_file_df["Functional Amount"])]

        # Second line updated
        copy_line_source91I_df = today_transaction_file_df.loc[(today_transaction_file_df["Journal Source"] == "91I") & (today_transaction_file_df["Functional Amount"])]
        copy_line_source91I_df["Functional Amount"] = ""
        copy_line_source91I_df["Daily Rec Comment"] = "RHA"
        copy_line_source91I_df["Master File Status"] = "RHA"
        copy_line_source91I_df["Transactional Amount"] = source_91I_difference_amount
        # Adding the RHA reject into the RHA sheet inside Master file
        rha_tobe_added_df = pd.concat([rha_tobe_added_df, difference_source91I_df], ignore_index=True)
    # --------------------------

    # Adding for source 91T
    if not difference_source91T_df.empty:
        # First line got updated
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91T") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91T") & today_transaction_file_df["Functional Amount"])]
        today_transaction_file_df.loc[((today_transaction_file_df["Journal Source"] == "91T") & today_transaction_file_df["Functional Amount"])]

        # Second line updated
        copy_line_source91T_df = today_transaction_file_df.loc[(today_transaction_file_df["Journal Source"] == "91T") & (today_transaction_file_df["Functional Amount"])]
        copy_line_source91T_df["Functional Amount"] = ""
        copy_line_source91T_df["Daily Rec Comment"] = "RHA"
        copy_line_source91T_df["Master File Status"] = "RHA"
        copy_line_source91T_df["Transactional Amount"] = source_91T_difference_amount
        # Adding the RHA reject into the RHA sheet inside Master file
        rha_tobe_added_df = pd.concat([rha_tobe_added_df, difference_source91T_df], ignore_index=True)
    # --------------------------

    # Adding all the extra columns to the rha_tobe_added dataframe
    if (not rha_tobe_added_df.empty):
        rha_tobe_added_df["MONTH"] = RHA_MONTH_DETAIL
        rha_tobe_added_df["OPENING BALANCE"] = ""
        rha_tobe_added_df["STATUS"] = "Open"
        rha_tobe_added_df["MONTH.1"] = RHA_MONTH_DETAIL
        rha_tobe_added_df["YEAR"] = int(YEAR)
        rha_tobe_added_df["CONCAT"] = ""
        rha_tobe_added_df["REVISED VALUE DATE"] = REVISED_VALUE_DATE
        rha_tobe_added_df["REVISED WORK OF DATE"] = REVISED_WORK_OF_DATE
    
    # --------------------------------------------------------------------------------------
    # This is step is executed after updating the transactional file dataframe

    master_file = MASTER_FILE_PATH
    master_file_df = pd.read_excel(master_file, sheet_name=None)

    # Updating the GL sheet
    logging.info(f"Updating the Master GL Sheet in the Master file with all the transaction file data and the RHA rejects.")
    transaction_file_sheet_df = master_file_df [MASTER_FILE_GL_SHEETNAME]
    aligned_today_transaction_file_df = today_transaction_file_df.reindex(columns=transaction_file_sheet_df.columns)
    updated_transaction_file_sheet_df = pd.concat([transaction_file_sheet_df, aligned_today_transaction_file_df], ignore_index=True) 
    master_file_df[MASTER_FILE_GL_SHEETNAME] = updated_transaction_file_sheet_df

    # Updating the RHA Sheet
    logging.info(f"Updating the RHA Sheet in the Master file with all the RHA rejects.")
    rha_master_sheet_df = master_file_df[MASTER_FILE_RHA_SHEETNAME]
    rha_master_sheet_df.columns = [col.upper() for col in rha_master_sheet_df.columns]

    aligned_rha_master_sheet_df = rha_tobe_added_df.reindex(columns=rha_master_sheet_df.columns)
    updated_rha_master_sheet_df = pd.concat([rha_master_sheet_df, aligned_rha_master_sheet_df], ignore_index=True)
    master_file_df[MASTER_FILE_RHA_SHEETNAME] = updated_rha_master_sheet_df

    with pd. ExcelWriter(master_file, engine='openpyxl') as writer:
        for sheet_name, data in master_file_df.items(): 
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    logging.info(f"GBP Values updation happened in the Master file successfully !!!")

    logging.info("Process completed for GBP Currency")
    logging.info("-" * 30)

    # --------------------------------------------------------------------------------------------

    logging.info("Processsing for Non-GBP Currency")

    logging.info("Retriving the Non-GBP data from Transactional file")
    today_transaction_file_df = transaction_file_df.loc[
        (transaction_file_df['Platform Proc Date']==PLATFORM_PROC_DATE) & 
        (transaction_file_df['Journal Source'] != 'GLR') &
        (transaction_file_df["Transactional Currency"] != 'GBP')
    ]

    today_transaction_file_df["Transactional Amount"] = today_transaction_file_df["Transactional Amount"].str.replace(",", "")
    today_transaction_file_df["Transactional Amount"] = today_transaction_file_df["Transactional Amount"].str.replace(")", "")
    today_transaction_file_df["Transactional Amount"] = today_transaction_file_df["Transactional Amount"].str.replace("(", "-").astype(float)

    today_transaction_file_df["Functional Amount"] = today_transaction_file_df["Functional Amount"].str.replace(",", "")
    today_transaction_file_df["Functional Amount"] = today_transaction_file_df["Functional Amount"].str.replace(")", "")
    today_transaction_file_df["Functional Amount"] = today_transaction_file_df["Functional Amount"].str.replace("(", "-").astype(float)

    today_transaction_file_df["Month"] = MONTH
    today_transaction_file_df["Year"] = int(YEAR)
    today_transaction_file_df["Business Unit"] = today_transaction_file_df["Business Unit"].astype(int)
    today_transaction_file_df["GL Accountt"] = today_transaction_file_df["GL Account"].astype(int)

    today_transaction_file_df["Revised Journal Source"] = today_transaction_file_df['Journal Source']
    today_transaction_file_df["Journal Source2"] = today_transaction_file_df['Journal Source']
    today_transaction_file_df["Daily Rec Comment"] = "RHA"
    today_transaction_file_df["Master File Status"] = "RHA"

    HUB_Source_list = ["91G", "91P", "91I", "91E", "91T", "91D"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Revised Journal Source'] = 'HUB'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Journal Source2'] = 'HUB'

    source_exclude_list = ["9RT"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Revised Journal Source'] = '9RT'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Journal Source2'] = '9RT'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Daily Rec Comment'] = 'Others'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Master File Status'] = 'Minor Diff'

    source_9AQ_list = ["9AQ"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Revised Journal Source'] = '9AQ'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Journal Source2'] = '9AQ'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Daily Rec Comment'] = 'Others'
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Master File Status'] = 'Cleared'

    source_P_list = ["PJE", "PJA", "MAN"]
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Daily Rec Comment'] = ''
    today_transaction_file_df.loc[today_transaction_file_df['Journal Source'].isin(HUB_Source_list), 'Master File Status'] = ''

    logging.info("Retriving the Non-GBP data from RHA file")
    rha_file = RHA_FILE_PATH
    rha_file_df = pd.read_excel(rha_file, sheet_name=None)
    rha_sheetname = list(rha_file_df.keys())[0]
    rha_extract_sheet_df = rha_file_df[rha_sheetname]
    rha_extract_sheet_df.columns = [col.upper() for col in rha_extract_sheet_df.columns]

    # Note: Loc statement should be checked                                                  <<<---------------------
    rha_extract_sheet_df = rha_extract_sheet_df.loc[
            (rha_extract_sheet_df['WORK OF DATE'] == RHA_WORK_OF_DATE) &
            (rha_extract_sheet_df['CURRENCY'] != CURRENCY) &
            (rha_extract_sheet_df['T/R'] == 'T') &
            (rha_extract_sheet_df['RPT UNIT'] == RHA_REPORTING_UNIT)
        ]

    rha_extract_sheet_df["MONTH"] = RHA_MONTH_DETAIL
    rha_extract_sheet_df["OPENING BALANCE"] = ""
    rha_extract_sheet_df["STATUS"] = "Open"
    rha_extract_sheet_df["MONTH.1"] = RHA_MONTH_DETAIL
    rha_extract_sheet_df["YEAR"] = int(YEAR)
    rha_extract_sheet_df["CONCAT"] = ""
    rha_extract_sheet_df["REVISED VALUE DATE"] = REVISED_VALUE_DATE
    rha_extract_sheet_df["REVISED WORK OF DATE"] = REVISED_WORK_OF_DATE

    # -------------------------------------------------------

    master_file = MASTER_FILE_PATH
    master_file_df = pd.read_excel(master_file, sheet_name=None)

    # Updating the GL sheet
    logging.info(f"Updating the Master GL Sheet in the Master file with Non-GBP data")
    transaction_file_sheet_df = master_file_df [MASTER_FILE_GL_SHEETNAME]
    aligned_today_transaction_file_df = today_transaction_file_df.reindex(columns=transaction_file_sheet_df.columns)
    updated_transaction_file_sheet_df = pd.concat([transaction_file_sheet_df, aligned_today_transaction_file_df], ignore_index=True) 
    master_file_df[MASTER_FILE_GL_SHEETNAME] = updated_transaction_file_sheet_df

    # Updating the RHA Sheet
    logging.info(f"Updating the RHA Sheet in the Master file with with Non-GBP data")
    rha_master_sheet_df = master_file_df[MASTER_FILE_RHA_SHEETNAME]
    rha_master_sheet_df.columns = [col.upper() for col in rha_master_sheet_df.columns]

    aligned_rha_master_sheet_df = rha_tobe_added_df.reindex(columns=rha_master_sheet_df.columns)
    updated_rha_master_sheet_df = pd.concat([rha_master_sheet_df, aligned_rha_master_sheet_df], ignore_index=True)
    master_file_df[MASTER_FILE_RHA_SHEETNAME] = updated_rha_master_sheet_df

    with pd. ExcelWriter(master_file, engine='openpyxl') as writer:
        for sheet_name, data in master_file_df.items(): 
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    master_file_df = pd.read_excel(master_file, sheet_name=None)
    wb = load_workbook(master_file)
    yellow_fill = PatternFill(start_color='E8FF00', end_color='E8FF00', fill_type='solid')

    for sheet_name, data in master_file_df.items():
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = yellow_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    wb.save(master_file)

    logging.info(f"Non-GBP values updation happened in the Master file successfully !!!")

    logging.info("Process completed for Non-GBP Currency")
    logging.info("-"*30)

    logging.info(f"SUCCESFULLY COMPLETED THE SFB PROCESS FOR {PLATFORM_PROC_DATE}")
    
    # ---------------------------------------------------------------------------------


except Exception as err:
    logging.error(f"An unexpected error occured: {err}")