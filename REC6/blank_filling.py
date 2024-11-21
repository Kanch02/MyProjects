import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings

warnings.filterwarnings("ignore")

try:
    print("PROCESS STARTED !!!")

    RHA_FILE_PATH = "...................\\RHA_filename.xlsx"    # <<----- Replace with your file name
    print(f"Reading the RHA file from path: {os.path.basename(RHA_FILE_PATH)}")

    rha_file_df = pd.read_excel(RHA_FILE_PATH, sheet_name=None)
    rha_sheetname = list(rha_file_df.keys())[0]
    rha_extract_sheet_df = rha_file_df[rha_sheetname]

    # Define the conditions and values for "ALN" column
    rha_extract_sheet_df.loc[(rha_extract_sheet_df['RPT UNIT'].isna()) & (rha_extract_sheet_df['ALN'].str.contains('HRBF', na=False)), 'RPT UNIT'] = "10080"
    rha_extract_sheet_df.loc[(rha_extract_sheet_df['RPT UNIT'].isna()) & (rha_extract_sheet_df['ALN'].str.contains('MDBK', na=False)), 'RPT UNIT'] = "22880"
    rha_extract_sheet_df.loc[(rha_extract_sheet_df['RPT UNIT'].isna()) & (rha_extract_sheet_df['ALN'].str.contains('888', na=False)), 'RPT UNIT'] = "HINV"

    # Define conditions for cases where "ALN" is blank and use "COST CENTRE" column
    rha_extract_sheet_df.loc[(rha_extract_sheet_df['RPT UNIT'].isna()) & (rha_extract_sheet_df['COST CENTRE'].str.startswith('4', na=False) | rha_extract_sheet_df['COST CENTRE'].str.startswith('6', na=False)), 'RPT UNIT'] = "22880"
    rha_extract_sheet_df.loc[(rha_extract_sheet_df['RPT UNIT'].isna()) & (rha_extract_sheet_df['COST CENTRE'].str.startswith('5', na=False) & ~rha_extract_sheet_df['COST CENTRE'].str.startswith('59', na=False)), 'RPT UNIT'] = "10080"
    rha_extract_sheet_df.loc[(rha_extract_sheet_df['RPT UNIT'].isna()) & (rha_extract_sheet_df['COST CENTRE'].str.startswith('59', na=False)), 'RPT UNIT'] = "11700"


    rha_file_df[rha_sheetname] = rha_extract_sheet_df

    print(f"Saving the details back to RHA file: {os.path.basename(RHA_FILE_PATH)}")
    with pd. ExcelWriter(RHA_FILE_PATH, engine='openpyxl') as writer:
        for sheet_name, data in rha_file_df.items(): 
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Formatting the RHA file !!!")
    rha_file_df = pd.read_excel(RHA_FILE_PATH, sheet_name=None)
    wb = load_workbook(rha_file_df)
    yellow_fill = PatternFill(start_color='E8FF00', end_color='E8FF00', fill_type='solid')

    for sheet_name, data in rha_file_df.items():
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = yellow_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    wb.save(rha_file_df)

    print("PROCESS END SUCCESSFULLY !!!")

except Exception as err:
    print(f"An unexpected error occured: {err}")













