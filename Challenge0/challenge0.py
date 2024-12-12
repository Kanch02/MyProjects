import os
import pandas as pd  # type: ignore
import numpy as np  # type: ignore
import yaml
import logging
import warnings
import openpyxl  # type: ignore
from datetime import datetime
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import PatternFill # type: ignore

warnings.filterwarnings("ignore")

# Creating the folders to store all the log files required 
# log_folder = ".\\Logs"
# os.makedirs(log_folder, exist_ok=True)
# log_file = log_folder + "\\CHALLENGE_0_" + datetime.now().strftime("%m%d%Y-%H%M%S") + ".log"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)-7s | %(message)s',
    handlers=[
        # logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)

logging.info(f"Loan formatting process started !!!")


# Retriving the values from config.yaml file
logging.info(f"Reading the config file: config.yaml")
with open("config.yaml", 'r') as file:
    config = yaml.safe_load(file)

LOAN_DATA_FILE_PATH = config['LOAN_DATA_FILE_PATH']
LOAN_DETAILS1_SHEET_NAME = config['LOAN_DETAILS1_SHEET_NAME']
LOAN_DETAILS2_SHEET_NAME = config['LOAN_DETAILS2_SHEET_NAME']
FX_RATES_SHEET_NAME = config['FX_RATES_SHEET_NAME']
OUTPUT_FILE_PATH = config['OUTPUT_FILE_PATH']
PAYMENT_FREQUENCY_MONTHLY =  int(config['PAYMENT_FREQUENCY_MONTHLY'])
PAYMENT_FREQUENCY_QUATERLY =  int(config['PAYMENT_FREQUENCY_QUATERLY'])
PAYMENT_FREQUENCY_HALF_YEARLY =  int(config['PAYMENT_FREQUENCY_HALF_YEARLY'])
PAYMENT_FREQUENCY_YEARLY =  int(config['PAYMENT_FREQUENCY_YEARLY'])

# Reading both the loan details sheet and gathering the information
loan_details_1_df = pd.read_excel(LOAN_DATA_FILE_PATH, sheet_name=LOAN_DETAILS1_SHEET_NAME)
loan_details_2_df = pd.read_excel(LOAN_DATA_FILE_PATH, sheet_name=LOAN_DETAILS2_SHEET_NAME)

# Reading the Forex exchange rate 
fx_rates_df = pd.read_excel(LOAN_DATA_FILE_PATH, sheet_name=FX_RATES_SHEET_NAME)
fx_rates_dict = {}
for index, row in fx_rates_df.iterrows():
    currency = row['Currency']
    rate = float(row['Rate'])
    fx_rates_dict[currency] = rate


# Combining both the details into a single conslidated one
logging.info(f"Combining both the loan details files into a single file.")
df_consolidated = pd.concat([loan_details_1_df, loan_details_2_df], ignore_index=True)

# Converting interest rates into percentage
df_consolidated['interest_rate'] = df_consolidated['interest_rate']*100

# Converting both the date columns into proper format
df_consolidated['maturity'] = pd.to_datetime(df_consolidated['maturity'], errors='coerce').dt.date
df_consolidated['reporting_date'] = pd.to_datetime(df_consolidated['reporting_date'], errors='coerce').dt.date

# Removing values which have face_value = 0
df_consolidated = df_consolidated[df_consolidated['face_value']!=0]

# Creating the face_value_USD column
df_consolidated['face_value_USD'] = df_consolidated.apply(
    lambda row: round(row['face_value']*fx_rates_dict.get(row['currency'],1), 2), axis=1
)

# Making the interest to be 6% for face_value_USD > 5,00,000
df_consolidated.loc[(df_consolidated['face_value_USD'] > 500000), 'interest_rate'] = 6

# Amortisation_type to Standard value if the value is none
df_consolidated['amortisation_type'] = df_consolidated['amortisation_type'].replace([None, ''], np.nan).fillna("Standard")

# Maturity date is empty, then it is assumed as next date of reporting_date
df_consolidated['maturity'] = df_consolidated['maturity'].fillna(df_consolidated['reporting_date'] + pd.Timedelta(days=1))

# Creating empty columns which are needed and later values will get added
df_consolidated['Payment Date'] = None
df_consolidated['ndays'] = None
df_consolidated['days from Reporting'] = None
df_consolidated['Interest_USD'] = None
df_consolidated['Installment_USD'] = None
df_consolidated['Bucket'] = None

# Dropping the source column
df_consolidated = df_consolidated.drop(columns=['Source'])

payment_frequency_dict = {
    "Monthly":PAYMENT_FREQUENCY_MONTHLY, 
    "Quaterly":PAYMENT_FREQUENCY_QUATERLY, 
    "HalfYearly":PAYMENT_FREQUENCY_HALF_YEARLY,
    "Yearly": PAYMENT_FREQUENCY_YEARLY
}

bucket_ranges = {
    '1m': 30,
    '1m to 3m': 90,
    '3m to 6m': 180,
    '6m to 9m': 270,
    '9m to 12m': 360,
    '1y to 1.5y': 540,
    '1.5y to 2y': 720,
    '2y to 3y': 1080,
    '3y to 5y': 1800,
    '5y to 10y': 3600,
    '10y to 15y': 5400,
    '15y to 20y': 7200,
    '20y+': float('inf')
}

# Function to get bucket based on days from reporting
def get_bucket(days):
    for bucket, max_days in bucket_ranges.items():
        if days <= max_days:
            return bucket
    return '20y+'

# This is the final dataframe which store the updated results
final_result = []

for index, row in df_consolidated.iterrows():
    reporting_date = row['reporting_date']
    maturity_date = row['maturity']
    nDays = payment_frequency_dict[row['payment_frequency']]
    last_payment_date = reporting_date

    # Processing until maturty date is reached
    while (last_payment_date <= maturity_date):
        payment_date = last_payment_date + pd.Timedelta(days=nDays)
        if(payment_date > maturity_date):
            break
        new_row = row.copy()
        new_row['Payment Date'] = payment_date
        new_row['ndays'] = nDays
        new_row['days from Reporting'] = (payment_date - reporting_date).days

        # Calculating the interest rate in USD
        interest_usd = (new_row['face_value_USD']*new_row['interest_rate']*new_row['ndays']) / (365 * 100)
        new_row['Interest_USD'] = round(interest_usd, 3)
        new_row['Installment_USD'] = new_row['Interest_USD']
        new_row['Bucket'] = get_bucket(int(new_row['days from Reporting']))

        final_result.append(new_row)

        last_payment_date = payment_date
        last_installment_usd = interest_usd + new_row['face_value_USD']

    final_result[-1]['Installment_USD'] = last_installment_usd

# Converting the result into dataframe
final_result_df = pd.DataFrame(final_result)

desired_column_order = ['transaction_id', 'amortisation_type', 'payment_frequency', 'currency', 'face_value', 'face_value_USD', 'interest_rate', 'maturity', 'reporting_date', 'Payment Date', 'ndays', 'days from Reporting', 'Interest_USD', 'Installment_USD', 'Bucket']
final_result_df = final_result_df[desired_column_order]

#--------------------------------------------------------------------------------------------

# Processing to calculate the bucket wise total
bucket_result = {
    'Row Labels': [],
    'Sum of Installment': []
}
bucket_sum = final_result_df.groupby('Bucket')['Installment_USD'].sum().reset_index()
bucket_result['Row Labels'] = bucket_sum['Bucket'].tolist()
bucket_result['Sum of Installment'] = bucket_sum['Installment_USD'].tolist()

grand_total = sum(bucket_result['Sum of Installment'])
bucket_result['Row Labels'].append('Grand Total') 
bucket_result['Sum of Installment'].append(grand_total)
bucket_df = pd.DataFrame(bucket_result)

#--------------------------------------------------------------------------------------------
# Saving the result into excel sheet
with pd.ExcelWriter(OUTPUT_FILE_PATH, engine='openpyxl', mode='w') as writer:
    final_result_df.to_excel(writer, index=False, sheet_name='Output1')
    bucket_df.to_excel(writer, index=False, sheet_name='Output2')

# Formatting the excel sheet
output_file_df = pd.read_excel(OUTPUT_FILE_PATH, sheet_name=None)
wb = load_workbook(OUTPUT_FILE_PATH)
colour_fill = PatternFill(start_color='A0A0A0', end_color='A0A0A0', fill_type='solid')

for sheet_name, data in output_file_df.items():
    ws = wb[sheet_name]
    for cell in ws[1]:
        cell.fill = colour_fill

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max_length + 1
        ws.column_dimensions[column].width = adjusted_width

wb.save(OUTPUT_FILE_PATH)

logging.info(f"Output saved at {OUTPUT_FILE_PATH} sucessfully !!!")